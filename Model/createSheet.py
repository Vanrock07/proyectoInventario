# Crea una hoja de Excel con encabezados personalizados y formato
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


workbook = Workbook()
sheet = workbook.active

# Encabezados
headers = [
    "Fecha", "Ciudad", "Sede", "Centro de costos","ubicacion", "Nombre de Usuario",
    "No de documento","Tipo de equipo", "Marca de computador", "Modelo", "Serial",
    "Activo","Tipo de disco","Tamano del disco","Memoria Ram", "Procesador","Sistema Operativo",
    "Marca de monitor", "Serial", "Activo", "Accesorios", "Observaciones"
]

for col, header in enumerate(headers, start=1):
    cell = sheet.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")  # Negrita y texto blanco
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")  # Centrar texto
    column_letter = cell.column_letter  # Ajustar ancho de columna al texto del encabezado
    sheet.column_dimensions[column_letter].width = len(header) + 2  # +2 para un poco de espacio extra# Fondo azul
    
sheet.auto_filter.ref = f"A1:V1" # Aplica autofiltro a los encabezados
workbook.save(filename="datosMantenimiento1.xlsx")

