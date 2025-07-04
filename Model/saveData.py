import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import load_workbook
import os


# workbook = openpyxl.load_workbook(excel_path)  # Cargar el archivo Excel
# sheet = workbook.active

class ExportData:
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self._validate_template_path() # validar la ruta de la plantilla Excel
        self.dataToExport = {} # Inicializar el diccionario de datos a exportar
        
    def _validate_template_path(self): # Validar la ruta de la plantilla Excel
            if not os.path.exists(self.excel_path):
                raise FileNotFoundError(f"Archivo Excel no encontrado: {self.excel_path}")  
            
               
    def saveDataToExcel(self, datos):
        
        self.dataToExport = datos  # Asignar los datos a exportar
        
        try:
            # Cargar el archivo Excel usando la ruta absoluta
            workbook = load_workbook(self.excel_path)
            sheet = workbook.active
            
            # Ejemplo: Escribir datos (personaliza según tu necesidad)
            row = 2 # La fila en la que empezar a escribir # La columna en la que empezar a escribir
            cell = sheet['A2']

            if cell.value:  #si hay dato en la celda inserte un fila vacia
                sheet.insert_rows(idx=2, amount=1) 
    
            for col, key in enumerate(self.dataToExport.keys(), start=1):    #iteracion sobre el diccionario 
                sheet.cell(row=row, column=col).value = self.dataToExport[key] # Escribe el valor en la celda indicada  
                sheet.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
                column_letter = sheet.cell(row=row, column=col).column_letter # Obtener la letra de la columna
                sheet.column_dimensions[column_letter].width = len(self.dataToExport) + 2 
                #column += 1
            
        
        # Guardar el archivo Excel modificado
            workbook.save(self.excel_path)  
        
        except Exception as e:
            print(f"Error al guardar los datos en Excel: {e}")
            raise e
        
            
           

